import io
import os
import openpyxl
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

SCOPES = ['https://www.googleapis.com/auth/drive']

def main():
    creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    drive_service = build('drive', 'v3', credentials=creds)

    file_id = '1_ohJciOqOJWzeBc_NAEb7vion-pBS65T'
    file_name = 'temp_edit.xlsx'

    print("Downloading file...")
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.FileIO(file_name, mode='wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    print("Downloaded.")

    print("Editing with openpyxl...")
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    
    # The user said they put 'x' in A700. Let's find it.
    found = False
    for row in range(690, 710):
        val = sheet.cell(row=row, column=1).value
        if val and str(val).strip().lower() == 'x':
            sheet.cell(row=row, column=1).value = 'eu estou aqui'
            print(f"Replaced 'x' with 'eu estou aqui' at A{row}!")
            found = True
            break
            
    if not found:
        print("Could not find 'x' near A700. Setting A700 to 'eu estou aqui'.")
        sheet['A700'] = 'eu estou aqui'
        
    wb.save(file_name)

    print("Uploading back to Google Drive...")
    media = MediaFileUpload(file_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    drive_service.files().update(fileId=file_id, media_body=media).execute()
    print("Done!")

if __name__ == '__main__':
    main()
