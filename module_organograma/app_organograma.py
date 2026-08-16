import os
import sys
import threading
import time
import json
from flask import Flask, send_from_directory, jsonify, request, make_response

current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.abspath(os.path.join(current_dir, '..'))
core_dir = os.path.join(root_dir, 'core')

if root_dir not in sys.path:
    sys.path.insert(0, root_dir)

from module_organograma.data_processor import (
    fetch_organograma_data,
    sync_configs_to_cloud,
    process_csv_files,
)

app = Flask(__name__, template_folder='ui', static_folder='ui')
PORT = 5009

# ─── versão ────────────────────────────────────────────────
def _read_version():
    try:
        vpath = os.path.join(core_dir, 'version.json')
        if os.path.exists(vpath):
            with open(vpath, 'r', encoding='utf-8') as f:
                return json.load(f).get('version', '—')
    except Exception:
        pass
    return '—'

APP_VERSION = _read_version()

# ─── API ────────────────────────────────────────────────────
@app.route('/api/status')
def api_status():
    return jsonify({"status": "DRIVE CONECTADO E PRONTO PARA DESENVOLVIMENTO", "ready": True, "version": _read_version()})

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/api/headcount')
def api_headcount():
    try:
        path = os.path.join(current_dir, 'data', 'headcount.json')
        if not os.path.exists(path):
            return jsonify([])
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        print(f"[Organograma] Erro ao ler headcount.json: {e}")
        return jsonify([])

@app.route('/api/afastamentos')
def api_afastamentos():
    try:
        path = os.path.join(current_dir, 'data', 'afastamentos.json')
        if not os.path.exists(path):
            return jsonify([])
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        print(f"[Organograma] Erro ao ler afastamentos.json: {e}")
        return jsonify([])

@app.route('/api/afastamentos_historico')
def api_afastamentos_historico():
    try:
        path = os.path.join(current_dir, 'data', 'afastamentos_historico.json')
        if not os.path.exists(path):
            return jsonify([])
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        print(f"[Organograma] Erro ao ler afastamentos_historico.json: {e}")
        return jsonify([])

@app.route('/api/storage/get/<path:key>')
def api_storage_get(key):
    safe_key = key.replace(':', '_').replace('/', '_')
    path = os.path.join(current_dir, 'data', f"{safe_key}.json")
    if not os.path.exists(path) and safe_key in ['organograma_contagem', 'organograma_posso_contar']:
        alt_key = 'organograma_posso_contar' if safe_key == 'organograma_contagem' else 'organograma_contagem'
        path = os.path.join(current_dir, 'data', f"{alt_key}.json")
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return jsonify({"value": f.read()})
    return jsonify({"value": None})

@app.route('/api/storage/set/<path:key>', methods=['POST'])
def api_storage_set(key):
    safe_key = key.replace(':', '_').replace('/', '_')
    os.makedirs(os.path.join(current_dir, 'data'), exist_ok=True)
    val = request.json.get('value', '')
    keys_to_write = [safe_key]
    if safe_key in ['organograma_contagem', 'organograma_posso_contar']:
        keys_to_write = ['organograma_contagem', 'organograma_posso_contar']
    for k in keys_to_write:
        path = os.path.join(current_dir, 'data', f"{k}.json")
        with open(path, 'w', encoding='utf-8') as f:
            f.write(val)
    # Async cloud sync so the UI response is instant
    threading.Thread(target=_safe_cloud_sync, daemon=True).start()
    return jsonify({"success": True})

def _safe_cloud_sync():
    try:
        sync_configs_to_cloud()
    except Exception as e:
        print(f"[Organograma] Cloud sync em background falhou: {e}")

@app.route('/api/refresh', methods=['POST'])
def api_refresh():
    """Endpoint acionado pelo botão 'Atualizar Informações'. Baixa CSVs do Drive e reprocessa."""
    def _do_refresh():
        try:
            print("[Organograma] Refresh manual iniciado pelo usuário...")
            fetch_organograma_data()
            print("[Organograma] Refresh manual concluído.")
        except Exception as e:
            print(f"[Organograma] Erro no refresh manual: {e}")

    t = threading.Thread(target=_do_refresh, daemon=True)
    t.start()
    t.join(timeout=60)   # aguarda até 60 s para responder
    return jsonify({"success": True, "message": "Dados atualizados com sucesso."})

@app.route('/api/metadata')
def api_metadata():
    try:
        path = os.path.join(current_dir, 'data', 'metadata.json')
        if not os.path.exists(path):
            return jsonify({})
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        print(f"[Organograma] Erro ao ler metadata.json: {e}")
        return jsonify({})

@app.route('/api/afastamentos_stats')
def api_afastamentos_stats():
    try:
        path = os.path.join(current_dir, 'data', 'afastamentos_stats.json')
        if not os.path.exists(path):
            return jsonify({})
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        print(f"[Organograma] Erro ao ler afastamentos_stats.json: {e}")
        return jsonify({})

@app.route('/api/version')
def api_version():
    return jsonify({"version": _read_version()})

# ─── UI injectors ───────────────────────────────────────────
def inject_navigation(html):
    nav = f"""
    <style>
      body {{ padding-left: 64px !important; transition: padding-left 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important; }}
      body:has(.master-sidebar.open) {{ padding-left: 280px !important; }}
      .master-sidebar {{
        position: fixed; top: 0; left: 0; bottom: 0; width: 64px;
        background: #182333; color: white; z-index: 10000;
        transition: width 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        display: flex; flex-direction: column; overflow: hidden;
        box-shadow: 2px 0 12px rgba(0,0,0,0.15); font-family: 'IBM Plex Sans', sans-serif;
        white-space: nowrap;
      }}
      .master-sidebar.open {{ width: 280px; }}
      
      .ms-header {{
        height: 64px; display: flex; align-items: center; padding: 0 20px;
        border-bottom: 1px solid rgba(255,255,255,0.05); cursor: pointer;
      }}
      .ms-header svg {{ flex-shrink: 0; width: 24px; height: 24px; color: #fff; opacity: 0.7; }}
      .ms-brand {{ 
        font-family: 'Space Grotesk', sans-serif; font-weight: 600; color: #fff; 
        font-size: 16px; margin-left: 20px; opacity: 0; transition: opacity 0.2s;
      }}
      .master-sidebar.open .ms-brand {{ opacity: 1; transition-delay: 0.1s;}}
      
      .ms-nav {{ display: flex; flex-direction: column; padding: 16px 0; flex: 1; }}
      .ms-item {{
        display: flex; align-items: center; padding: 12px 20px;
        color: #8A96A3; text-decoration: none; font-weight: 500; font-size: 14.5px;
        transition: color 0.2s, background 0.2s; border-left: 3px solid transparent;
      }}
      .ms-item:hover {{ color: #FFFFFF; background: rgba(255,255,255,0.05); }}
      .ms-item.active {{ color: #FFFFFF; background: rgba(36,80,124,0.4); border-left-color: #5C95C6; }}
      .ms-item svg {{ flex-shrink: 0; width: 22px; height: 22px; margin-right: 22px; margin-left: 1px; }}
      .ms-label {{ opacity: 0; transition: opacity 0.2s; }}
      .master-sidebar.open .ms-label {{ opacity: 1; transition-delay: 0.1s; }}

      /* Version badge at bottom */
      .ms-version {{
        padding: 12px 20px; border-top: 1px solid rgba(255,255,255,0.05);
        font-size: 10.5px; color: rgba(255,255,255,0.28); font-family: 'IBM Plex Mono', monospace;
        white-space: nowrap; overflow: hidden;
        opacity: 0; transition: opacity 0.2s;
      }}
      .master-sidebar.open .ms-version {{ opacity: 1; transition-delay: 0.1s; }}

      /* Refresh button */
      .ms-refresh {{
        display: flex; align-items: center; margin: 8px 12px; padding: 9px 12px;
        background: rgba(36,80,124,0.3); border: 1px solid rgba(92,149,198,0.25);
        border-radius: 8px; color: #8A96A3; cursor: pointer; gap: 10px;
        font-size: 13px; font-family: 'IBM Plex Sans', sans-serif; font-weight: 500;
        transition: background 0.2s, color 0.2s, opacity 0.2s;
        opacity: 0; pointer-events: none;
        white-space: nowrap;
      }}
      .master-sidebar.open .ms-refresh {{ opacity: 1; pointer-events: auto; transition-delay: 0.1s; }}
      .ms-refresh:hover {{ background: rgba(36,80,124,0.6); color: #fff; }}
      .ms-refresh:disabled {{ opacity: 0.4 !important; cursor: not-allowed; }}
      .ms-refresh svg {{ flex-shrink: 0; width: 16px; height: 16px; }}
      @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
      .ms-refresh.loading svg {{ animation: spin 1s linear infinite; }}
    </style>
    <div class="master-sidebar" id="masterSidebar">
      <div class="ms-header" id="msHeader">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="3" y1="12" x2="21" y2="12"/><line x1="3" y1="6" x2="21" y2="6"/><line x1="3" y1="18" x2="21" y2="18"/></svg>
        <span class="ms-brand">Organograma Tool</span>
      </div>
      <div class="ms-nav">
        <a href="/" class="ms-item">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path><circle cx="9" cy="7" r="4"></circle><path d="M23 21v-2a4 4 0 0 0-3-3.87"></path><path d="M16 3.13a4 4 0 0 1 0 7.75"></path></svg>
          <span class="ms-label">Cadastro de Responsáveis</span>
        </a>
        <a href="/headcount" class="ms-item">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"/><line x1="3" y1="9" x2="21" y2="9"/><line x1="9" y1="21" x2="9" y2="9"/></svg>
          <span class="ms-label">Headcount / Auditoria</span>
        </a>
        <a href="/arvore" class="ms-item">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7"></rect><rect x="14" y="3" width="7" height="7"></rect><rect x="14" y="14" width="7" height="7"></rect><rect x="3" y="14" width="7" height="7"></rect></svg>
          <span class="ms-label">Visão Organograma</span>
        </a>
        <a href="/sugestao" class="ms-item">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path><polyline points="14 2 14 8 20 8"></polyline><line x1="12" y1="18" x2="12" y2="12"></line><line x1="9" y1="15" x2="15" y2="15"></line></svg>
          <span class="ms-label">Sugestão de Organograma</span>
        </a>
      </div>
      <button class="ms-refresh" id="msRefreshBtn">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="23 4 23 10 17 10"/><polyline points="1 20 1 14 7 14"/><path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"/></svg>
        <span class="ms-refresh-label">Atualizar Informações</span>
      </button>
      <div class="ms-version">v{_read_version()} · Organograma Tool</div>
    </div>
    <script>
      document.querySelectorAll('.ms-nav .ms-item').forEach(link => {{
        if (link.getAttribute('href') === window.location.pathname) link.classList.add('active');
      }});
      document.getElementById('msHeader').addEventListener('click', () => {{
        document.getElementById('masterSidebar').classList.toggle('open');
      }});

      // Refresh button handler
      document.getElementById('msRefreshBtn').addEventListener('click', async function() {{
        const btn = this;
        const label = btn.querySelector('.ms-refresh-label');
        btn.disabled = true;
        btn.classList.add('loading');
        label.textContent = 'Atualizando...';
        try {{
          const res = await fetch('/api/refresh', {{ method: 'POST' }});
          const data = await res.json();
          label.textContent = data.success ? 'Atualizado!' : 'Erro';
          setTimeout(() => {{
            label.textContent = 'Atualizar Informações';
            btn.disabled = false;
            btn.classList.remove('loading');
            // Reload page to show fresh data
            window.location.reload();
          }}, 1800);
        }} catch(e) {{
          label.textContent = 'Erro de conexão';
          setTimeout(() => {{
            label.textContent = 'Atualizar Informações';
            btn.disabled = false;
            btn.classList.remove('loading');
          }}, 2000);
        }}
      }});
    </script>
    """
    return html.replace('<body>', '<body>' + nav)

def inject_storage(html):
    storage_script = """
    <script>
    window.storage = {
        get: async function(key, fallback) {
            try {
                const res = await fetch('/api/storage/get/' + encodeURIComponent(key));
                const data = await res.json();
                return data.value ? {value: data.value} : null;
            } catch(e) { return null; }
        },
        set: async function(key, value, sync) {
            try {
                const res = await fetch('/api/storage/set/' + encodeURIComponent(key), {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({value: value})
                });
                return await res.json();
            } catch(e) { return false; }
        }
    };
    </script>
    """
    if '</body>' in html:
        return html.replace('</body>', storage_script + '</body>')
    return html + storage_script

def _render_page(html_file):
    path = os.path.join(current_dir, 'ui', html_file)
    if not os.path.exists(path):
        return f"Arquivo UI não encontrado: {html_file}"
    with open(path, 'r', encoding='utf-8') as f:
        html = f.read()
    html = inject_navigation(html)
    html = inject_storage(html)
    response = make_response(html)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/')
def index():
    return _render_page('index.html')

@app.route('/headcount')
def route_headcount():
    return _render_page('headcount.html')

@app.route('/arvore')
def route_arvore():
    return _render_page('organograma.html')

@app.route('/sugestao')
def route_sugestao():
    import os
    from flask import make_response, request
    path = os.path.join(current_dir, 'Template_Sugestao_Excel.html')
    if not os.path.exists(path) or request.args.get('refresh') == '1':
        try:
            from module_organograma.gen_template import generate_template
            generate_template()
        except Exception:
            try:
                from gen_template import generate_template
                generate_template()
            except Exception as e:
                print(f"[Organograma] Erro ao auto-gerar sugestão: {e}")
                
    if not os.path.exists(path):
        return "Arquivo de Sugestão não gerado ainda."
    with open(path, 'r', encoding='utf-8') as f:
        html = f.read()
    html = inject_navigation(html)
    response = make_response(html)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

@app.route('/api/sugestao/gerar', methods=['POST', 'GET'])
def api_sugestao_gerar():
    try:
        try:
            from module_organograma.gen_template import generate_template
            generate_template()
        except Exception:
            from gen_template import generate_template
            generate_template()
        return jsonify({"success": True, "message": "Template de sugestão gerado com sucesso."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/organograma_excel')
def api_export_organograma_excel():
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        from flask import send_file

        hc_path = os.path.join(current_dir, 'data', 'headcount.json')
        headcount = []
        if os.path.exists(hc_path):
            with open(hc_path, 'r', encoding='utf-8') as f:
                headcount = json.load(f)

        resp_path = os.path.join(current_dir, 'data', 'organograma_responsaveis.json')
        responsaveis = {}
        if os.path.exists(resp_path):
            with open(resp_path, 'r', encoding='utf-8') as f:
                responsaveis = json.load(f)
                for cc in responsaveis:
                    if not isinstance(responsaveis[cc], list):
                        arr = []
                        roleLabels = {'diretor':'Diretor', 'gerente':'Gerente', 'gerente2':'Gerente', 'coordenador':'Coordenador', 'coordenador2':'Coordenador', 'lider':'Líder', 'lider2':'Líder'}
                        for k, r in responsaveis[cc].items():
                            if r:
                                r['label'] = roleLabels.get(k, r.get('roleLabel', 'Gestor'))
                                arr.append(r)
                        responsaveis[cc] = arr

        cont_path = os.path.join(current_dir, 'data', 'organograma_posso_contar.json')
        if not os.path.exists(cont_path):
            cont_path = os.path.join(current_dir, 'data', 'organograma_contagem.json')
        posso_contar = {}
        if os.path.exists(cont_path):
            with open(cont_path, 'r', encoding='utf-8') as f:
                try:
                    raw_cont = f.read().strip()
                    if raw_cont:
                        posso_contar = json.loads(raw_cont)
                except Exception as e:
                    print(f"[Organograma] Erro ao ler contagem: {e}")

        af_stats_path = os.path.join(current_dir, 'data', 'afastamentos_stats.json')
        afast_stats = {}
        if os.path.exists(af_stats_path):
            with open(af_stats_path, 'r', encoding='utf-8') as f:
                afast_stats = json.load(f)

        cc_map = {}
        for e in headcount:
            cc_cod = e.get('ccCod')
            if cc_cod and cc_cod not in cc_map:
                cc_map[cc_cod] = {'cod': cc_cod, 'nome': e.get('ccNome', cc_cod)}
        
        cost_centers = sorted(list(cc_map.values()), key=lambda x: x['nome'])

        wb = openpyxl.Workbook()

        # Sheet 1: Organograma Geral
        ws_colab = wb.active
        ws_colab.title = "Organograma Geral"
        
        # Sheet 2: Resumo por CC
        ws_cc = wb.create_sheet(title="Resumo por CC")
        
        # Sheet 3: Linha de Gestão
        ws_gestao = wb.create_sheet(title="Linha de Gestão")

        header_fill = PatternFill(start_color="182333", end_color="182333", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_header_fill = PatternFill(start_color="24507C", end_color="24507C", fill_type="solid")
        sub_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin', color='D7DEE4'),
            right=Side(style='thin', color='D7DEE4'),
            top=Side(style='thin', color='D7DEE4'),
            bottom=Side(style='thin', color='D7DEE4')
        )

        headers_colab = [
            "Diretoria", "Gerência", "Coordenação", "Liderança",
            "Cód. CC", "Centro de Custo", "CAD", "Colaborador",
            "Cargo", "Tipo", "Admissão", "Faltas", "Atestados", "Acidentes", "Salário (R$)"
        ]
        ws_colab.append(headers_colab)
        for col_num, h in enumerate(headers_colab, 1):
            cell = ws_colab.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        headers_cc = [
            "Diretoria", "Cód. CC", "Centro de Custo",
            "Linha de Gestão", "Headcount", "Custo Total (R$)", "Média Salarial (R$)"
        ]
        ws_cc.append(headers_cc)
        for col_num, h in enumerate(headers_cc, 1):
            cell = ws_cc.cell(row=1, column=col_num)
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        headers_gestao = [
            "Diretoria", "Nível / Função", "CAD", "Nome do Gestor",
            "Tipo", "Salário (R$)", "Centros de Custo Atendidos"
        ]
        ws_gestao.append(headers_gestao)
        for col_num, h in enumerate(headers_gestao, 1):
            cell = ws_gestao.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        active_hc = [e for e in headcount if posso_contar.get(e.get('cad'), True) is not False]

        processed_gestores = set()

        for cc in cost_centers:
            cc_cod = cc['cod']
            cc_name = cc['nome']
            resp = responsaveis.get(cc_cod, [])
            base_hc = [e for e in active_hc if e.get('ccCod') == cc_cod]

            if not base_hc and not resp:
                continue

            dir_name = "Sem Diretoria Atribuída"
            ger_name = "—"
            coord_name = "—"
            lider_name = "—"

            for r in resp:
                lbl = r.get('label', '')
                if lbl == 'Diretor' and r.get('nome'):
                    dir_name = r['nome']
                elif lbl == 'Gerente' and r.get('nome'):
                    ger_name = r['nome']
                elif lbl == 'Coordenador' and r.get('nome'):
                    coord_name = r['nome']
                elif lbl == 'Líder' and r.get('nome'):
                    lider_name = r['nome']

            gestores_str = " | ".join([f"{r.get('label', 'Gestor')}: {r.get('nome', '')}" for r in resp if r.get('nome')])

            # Sheet 3: Gestores
            for r in resp:
                if not r or not r.get('nome'):
                    continue
                g_key = f"{r.get('cad') or r.get('nome')}_{r.get('label')}"
                if g_key not in processed_gestores:
                    processed_gestores.add(g_key)
                    ccs_managed = []
                    for cod, res_list in responsaveis.items():
                        if any(x.get('nome') == r.get('nome') for x in res_list if isinstance(x, dict)):
                            c_obj = next((c for c in cost_centers if c['cod'] == cod), None)
                            ccs_managed.append(c_obj['nome'] if c_obj else cod)
                    ccs_str = ", ".join(sorted(list(set(ccs_managed))))

                    row_g = [
                        dir_name,
                        r.get('label', 'Gestor'),
                        r.get('cad', '—'),
                        r.get('nome'),
                        r.get('tipo', '—'),
                        float(r.get('salario') or 0),
                        ccs_str
                    ]
                    ws_gestao.append(row_g)

            # Sheet 1: Colaboradores
            cc_total_sal = 0.0
            cc_count = 0
            for e in base_hc:
                cc_count += 1
                sal = float(e.get('salario') or 0)
                cc_total_sal += sal

                cad = e.get('cad', '')
                st = afast_stats.get(cad, {'faltas': 0, 'atestado': 0, 'acidente': 0})

                row_c = [
                    dir_name,
                    ger_name,
                    coord_name,
                    lider_name,
                    cc_cod,
                    cc_name,
                    cad or '—',
                    e.get('nome', '—'),
                    e.get('cargo', 'Sem Função'),
                    e.get('tipo', 'CLT'),
                    e.get('admissao', '—'),
                    int(st.get('faltas', 0)),
                    int(st.get('atestado', 0)),
                    int(st.get('acidente', 0)),
                    sal
                ]
                ws_colab.append(row_c)

            # Sheet 2: Resumo por CC
            if cc_count > 0 or resp:
                avg_sal = (cc_total_sal / cc_count) if cc_count > 0 else 0.0
                row_cc = [
                    dir_name,
                    cc_cod,
                    cc_name,
                    gestores_str or 'Sem Gestor Direto',
                    cc_count,
                    cc_total_sal,
                    avg_sal
                ]
                ws_cc.append(row_cc)

        for sheet in [ws_colab, ws_cc, ws_gestao]:
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        h_val = str(sheet.cell(row=1, column=cell.column).value or '')
                        if 'Salário' in h_val or 'Custo' in h_val or 'Média' in h_val:
                            cell.number_format = '#,##0.00'

            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Organograma_Tela3_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"[Organograma] Erro ao gerar Excel: {e}")
        return jsonify({"error": str(e)}), 500


# ─── background autopilot ───────────────────────────────────
def vigia_autopilot():
    print("[Organograma Vigia] Iniciando thread Autopilot...")
    while True:
        try:
            print("[Organograma Vigia] Sincronizando dados com o Google Drive (Autopilot)...")
            fetch_organograma_data()
        except Exception as e:
            print(f"[Organograma Vigia] Erro no autopilot: {e}")
        time.sleep(300)

def main():
    print(f"[Organograma] Bootloader v{_read_version()} — Iniciando módulo Organograma.")
    vigia_thread = threading.Thread(target=vigia_autopilot, daemon=True)
    vigia_thread.start()
    try:
        app.run(host='0.0.0.0', port=PORT, debug=False, use_reloader=False)
    except Exception as e:
        print(f"[Organograma] Erro ao iniciar servidor na porta {PORT}: {e}")


if __name__ == '__main__':
    main()
